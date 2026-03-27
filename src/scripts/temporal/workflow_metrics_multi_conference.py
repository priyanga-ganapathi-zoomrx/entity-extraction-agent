#!/usr/bin/env python3
"""Multi-conference workflow metrics exporter.

Builds a single XLSX workbook with:
- One sheet per conference (abstract-level workflow metrics rows)
- One analysis sheet with token rollups:
  1) conference + model level
  2) model level across conferences

Conference discovery is performed from a GCS root path and ignores known
non-conference folders (congress, prompts, rules).
"""

import argparse
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from tqdm import tqdm

from src.agents.core.storage import GCSStorageClient, parse_gcs_path
from src.scripts.temporal.utils import get_data_storage, list_abstract_ids, load_status
from src.scripts.temporal.workflow_metrics import STEPS, get_fieldnames, status_to_row


DEFAULT_EXCLUDED_FOLDERS = {"congress", "prompts", "rules"}

# Step path in status.json -> env var key
STEP_TO_ENV_VAR = {
    "drug.extraction": "DRUG_EXTRACTION_MODEL",
    "drug.validation": "DRUG_VALIDATION_MODEL",
    "drug_class.step1_regimen": "DRUG_CLASS_REGIMEN_MODEL",
    "drug_class.step2_extraction": "DRUG_CLASS_EXTRACTION_MODEL",
    "drug_class.step3_selection": "DRUG_CLASS_SELECTION_MODEL",
    "drug_class.step4_explicit": "DRUG_CLASS_EXPLICIT_MODEL",
    "drug_class.step5_consolidation": "DRUG_CLASS_CONSOLIDATION_MODEL",
    "drug_class.validation": "DRUG_CLASS_VALIDATION_MODEL",
    "indication.extraction": "INDICATION_LLM_MODEL",
    "indication.validation": "INDICATION_VALIDATION_LLM_MODEL",
}

# Env var key -> concrete model value (as requested)
ENV_MODEL_VALUES = {
    "DRUG_CLASS_REGIMEN_MODEL": "anthropic/claude-sonnet-4-6",
    "DRUG_CLASS_EXTRACTION_MODEL": "anthropic/claude-sonnet-4-6",
    "DRUG_CLASS_GROUNDED_MODEL": "gemini-3-flash-preview",
    "DRUG_CLASS_SELECTION_MODEL": "gemini-3-flash-preview",
    "DRUG_CLASS_EXPLICIT_MODEL": "gemini-3-flash-preview",
    "DRUG_CLASS_CONSOLIDATION_MODEL": "gemini-3-flash-preview",
    "DRUG_CLASS_VALIDATION_MODEL": "gemini-3-flash-preview",
    "DRUG_EXTRACTION_MODEL": "gemini-2.5-pro",
    "DRUG_VALIDATION_MODEL": "gemini-3-flash-preview",
    "INDICATION_LLM_MODEL": "gemini-2.5-pro",
    "INDICATION_VALIDATION_LLM_MODEL": "claude-sonnet-4-5-20250929",
}

_CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_SHEET_INVALID_CHARS_RE = re.compile(r"[\[\]\*\?/\\:]")


@dataclass
class ConferenceInfo:
    """Metadata for a discovered conference folder."""

    name: str
    prefix: str
    run_timestamp: str

    @property
    def run_date(self) -> str:
        return self.run_timestamp[:10] if self.run_timestamp else "unknown-date"


def _sanitize_cell(value) -> str:
    if value is None:
        return ""
    text = str(value)
    return _CONTROL_CHAR_RE.sub("", text)


def _build_sheet_name(base_name: str, used_names: set[str]) -> str:
    """Create Excel-safe unique sheet name with <=31 chars."""
    safe = _SHEET_INVALID_CHARS_RE.sub("_", base_name.strip())
    safe = safe or "sheet"
    safe = safe[:31]
    if safe not in used_names:
        used_names.add(safe)
        return safe

    # Add numeric suffix while preserving 31-char limit.
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = safe[: 31 - len(suffix)] + suffix
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate

    # Extremely unlikely fallback.
    fallback = f"sheet_{len(used_names) + 1}"
    fallback = fallback[:31]
    used_names.add(fallback)
    return fallback


def _get_nested(data: dict, dotted_path: str) -> dict:
    """Navigate nested dict by dotted path and return dict or {}."""
    current = data
    for key in dotted_path.split("."):
        if not isinstance(current, dict):
            return {}
        current = current.get(key, {})
    return current if isinstance(current, dict) else {}


def _discover_conferences(gcs_root: str) -> list[ConferenceInfo]:
    """Discover valid conference folders under GCS root."""
    bucket_name, root_prefix = parse_gcs_path(gcs_root)
    root_prefix = root_prefix.strip("/")
    listing_prefix = f"{root_prefix}/" if root_prefix else ""

    gcs = GCSStorageClient(bucket_name, "")
    bucket = gcs.bucket

    iterator = bucket.list_blobs(prefix=listing_prefix, delimiter="/")
    _ = list(iterator)  # consume iterator so prefixes are populated

    conference_infos: list[ConferenceInfo] = []
    for prefix_path in sorted(iterator.prefixes):
        # Extract folder name from "<root>/ConferenceName/"
        folder_name = prefix_path.rstrip("/").split("/")[-1]
        if not folder_name:
            continue
        if folder_name.lower() in DEFAULT_EXCLUDED_FOLDERS:
            continue

        conference_prefix = (
            f"{listing_prefix}{folder_name}" if listing_prefix else folder_name
        )
        abstract_titles_blob_path = f"{conference_prefix}/abstract_titles.csv"

        abstract_titles_blob = bucket.get_blob(abstract_titles_blob_path)
        if abstract_titles_blob is None:
            continue

        # Ensure abstracts path has at least one object.
        abstracts_prefix = f"{conference_prefix}/abstracts/"
        abstracts_iter = bucket.list_blobs(prefix=abstracts_prefix, max_results=1)
        has_abstracts = any(True for _ in abstracts_iter)
        if not has_abstracts:
            continue

        run_ts = (
            abstract_titles_blob.updated.isoformat()
            if abstract_titles_blob.updated
            else ""
        )
        conference_infos.append(
            ConferenceInfo(
                name=folder_name,
                prefix=conference_prefix,
                run_timestamp=run_ts,
            )
        )

    return conference_infos


def _save_workbook(workbook: Workbook, output_path: str) -> None:
    """Save workbook to local path or GCS."""
    if output_path.startswith("gs://"):
        bucket_name, full_prefix = parse_gcs_path(output_path)
        if "/" in full_prefix:
            base_prefix = "/".join(full_prefix.split("/")[:-1])
            filename = full_prefix.split("/")[-1]
        else:
            base_prefix = ""
            filename = full_prefix

        output_storage = GCSStorageClient(bucket_name, base_prefix)
        buffer = io.BytesIO()
        workbook.save(buffer)
        output_storage.upload_bytes(
            filename,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        print(f"\nXLSX file saved to GCS: {output_path}")
        return

    workbook.save(output_path)
    print(f"\nXLSX file saved: {output_path}")


def _append_analysis_sheet(
    workbook: Workbook,
    conference_model_agg: dict[tuple[str, str], dict],
    overall_model_agg: dict[str, dict],
) -> None:
    """Add a single analysis sheet with both aggregation views."""
    ws = workbook.create_sheet(title="analysis")

    ws.append(["Conference + Model Token Consumption"])
    ws.append(
        [
            "conference",
            "conference_run_date",
            "model",
            "input_tokens",
            "output_tokens",
            "total_tokens",
            "llm_calls",
            "abstract_count",
        ]
    )

    for key in sorted(conference_model_agg.keys()):
        data = conference_model_agg[key]
        ws.append(
            [
                data["conference"],
                data["conference_run_date"],
                data["model"],
                data["input_tokens"],
                data["output_tokens"],
                data["total_tokens"],
                data["llm_calls"],
                len(data["abstract_ids"]),
            ]
        )

    ws.append([])
    ws.append(["Model Token Consumption Across Conferences"])
    ws.append(
        [
            "model",
            "input_tokens",
            "output_tokens",
            "total_tokens",
            "llm_calls",
            "conference_count",
            "abstract_count",
        ]
    )

    for model in sorted(overall_model_agg.keys()):
        data = overall_model_agg[model]
        ws.append(
            [
                data["model"],
                data["input_tokens"],
                data["output_tokens"],
                data["total_tokens"],
                data["llm_calls"],
                len(data["conference_names"]),
                len(data["abstract_ids"]),
            ]
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export workflow metrics across conferences into one XLSX"
    )
    parser.add_argument(
        "--gcs_root",
        required=True,
        help="GCS root containing conference folders (e.g. gs://bucket)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output XLSX file path (local or gs://). "
        "Default: workflow_metrics_multi_conference_<timestamp>.xlsx",
    )
    args = parser.parse_args()

    if not args.output:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        args.output = f"workflow_metrics_multi_conference_{timestamp}.xlsx"

    print("Multi-Conference Workflow Metrics Exporter")
    print("=" * 72)
    print(f"GCS root: {args.gcs_root}")
    print(f"Output:   {args.output}")
    print()

    conferences = _discover_conferences(args.gcs_root)
    print(f"Discovered {len(conferences)} conference folders")
    if not conferences:
        print("No valid conference folders found.")
        return

    workbook = Workbook(write_only=True)
    # Remove default sheet from write-only workbook naming behavior by not using it.
    used_sheet_names: set[str] = set()

    fieldnames = get_fieldnames()

    conference_model_agg: dict[tuple[str, str], dict] = {}
    overall_model_agg: dict[str, dict] = {}

    bucket_name, _ = parse_gcs_path(args.gcs_root)
    total_status_files = 0
    for conference in conferences:
        conference_path = f"gs://{bucket_name}/{conference.prefix}"
        storage = get_data_storage(conference_path)
        abstract_ids = list_abstract_ids(conference_path)

        print(
            f"\n[{conference.name}] run_date={conference.run_date} "
            f"abstracts={len(abstract_ids)}"
        )
        if not abstract_ids:
            continue

        sheet_label = f"{conference.name}_{conference.run_date}"
        sheet_name = _build_sheet_name(sheet_label, used_sheet_names)
        ws = workbook.create_sheet(title=sheet_name)
        ws.append(fieldnames)

        for abstract_id in tqdm(
            abstract_ids,
            desc=f"{conference.name} status",
            unit="abstract",
            leave=False,
        ):
            status = load_status(storage, abstract_id)
            if not status:
                continue

            total_status_files += 1
            row = status_to_row(status)
            ws.append([_sanitize_cell(row.get(col, "")) for col in fieldnames])

            # Aggregate tokens by conference + model and overall model.
            for step_path, _ in STEPS:
                env_key = STEP_TO_ENV_VAR.get(step_path)
                if not env_key:
                    continue
                model_name = ENV_MODEL_VALUES.get(env_key, env_key)

                step_data = _get_nested(status, step_path)
                input_tokens = int(step_data.get("input_tokens", 0) or 0)
                output_tokens = int(step_data.get("output_tokens", 0) or 0)
                total_tokens = int(step_data.get("tokens", 0) or 0)
                llm_calls = int(step_data.get("llm_calls", 0) or 0)

                if (
                    input_tokens == 0
                    and output_tokens == 0
                    and total_tokens == 0
                    and llm_calls == 0
                ):
                    continue

                conf_model_key = (conference.name, model_name)
                if conf_model_key not in conference_model_agg:
                    conference_model_agg[conf_model_key] = {
                        "conference": conference.name,
                        "conference_run_date": conference.run_date,
                        "model": model_name,
                        "input_tokens": 0,
                        "output_tokens": 0,
                        "total_tokens": 0,
                        "llm_calls": 0,
                        "abstract_ids": set(),
                    }
                c_row = conference_model_agg[conf_model_key]
                c_row["input_tokens"] += input_tokens
                c_row["output_tokens"] += output_tokens
                c_row["total_tokens"] += total_tokens
                c_row["llm_calls"] += llm_calls
                c_row["abstract_ids"].add(abstract_id)

                if model_name not in overall_model_agg:
                    overall_model_agg[model_name] = {
                        "model": model_name,
                        "input_tokens": 0,
                        "output_tokens": 0,
                        "total_tokens": 0,
                        "llm_calls": 0,
                        "conference_names": set(),
                        "abstract_ids": set(),
                    }
                m_row = overall_model_agg[model_name]
                m_row["input_tokens"] += input_tokens
                m_row["output_tokens"] += output_tokens
                m_row["total_tokens"] += total_tokens
                m_row["llm_calls"] += llm_calls
                m_row["conference_names"].add(conference.name)
                m_row["abstract_ids"].add(abstract_id)

    _append_analysis_sheet(workbook, conference_model_agg, overall_model_agg)
    _save_workbook(workbook, args.output)

    print("\nDone.")
    print(f"- Conferences processed: {len(conferences)}")
    print(f"- status.json files loaded: {total_status_files}")
    print(f"- Analysis rows (conference+model): {len(conference_model_agg)}")
    print(f"- Analysis rows (overall model): {len(overall_model_agg)}")


if __name__ == "__main__":
    main()
